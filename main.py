from __future__ import annotations

import csv
import json
import re
import unicodedata
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
from fastapi import FastAPI, File, HTTPException, UploadFile
from fastapi.responses import HTMLResponse, JSONResponse
from fastapi.staticfiles import StaticFiles


BASE_DIR = Path(__file__).resolve().parent
STATIC_DIR = BASE_DIR / "static"
STORAGE_DIR = BASE_DIR / "storage"
STATE_FILE = STORAGE_DIR / "portfolio.json"

app = FastAPI(title="Consolidador de Carteira", version="1.0.0")
app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")

KNOWN_CLASSES = {"RF", "RV", "FUNDOS", "OUTROS"}
FUND_HINTS = {"fii", "fundo", "infra", "imobili", "multimerc", "credito", "crédito", "fip", "fiagro"}
RF_TOKENS = {"cdb", "cra", "cri", "lca", "lf", "deb", "debenture", "tpf", "ntn-b", "ntnb", "ntn-b1", "ntnb1"}
INDEXER_MAP = {"cdi": "CDI", "ipca": "IPCA", "pre": "Pré", "prefix": "Pré", "prefixado": "Pré"}
INSTITUTIONS = {
    "xp": "XP",
    "safra": "Safra",
    "btg": "BTG",
    "itau": "Itaú",
    "itaú": "Itaú",
    "bb": "BB",
    "brb": "BRB",
    "bmg": "BMG",
    "picpay": "PicPay",
    "original": "Original",
    "facta": "Facta Financeira",
    "pine": "Pine",
    "semear": "Semear",
    "bocom bbm": "Bocom BBM",
    "equatorial": "Equatorial Goiás",
    "iguá": "Iguá",
    "iguatemi": "Iguatemi",
    "rabobank": "Rabobank",
    "digimais": "Digimais",
    "neon": "Neon Financeira",
    "lebes": "Lebes Financeira",
    "arbi": "Banco Arbi",
    "nordeste": "Banco do Nordeste",
    "bdmg": "BDMG",
    "sabesp": "Sabesp",
    "einstein": "Albert Einstein",
}


def _ensure_storage() -> None:
    STORAGE_DIR.mkdir(exist_ok=True)


def _slug(v: Any) -> str:
    text = unicodedata.normalize("NFKD", str(v or ""))
    text = text.encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]+", " ", text.lower()).strip()


def _txt(v: Any) -> str:
    return str(v).strip() if v not in (None, "") else ""


def _money(v: Any) -> float | None:
    if v in (None, ""):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = _txt(v).replace("R$", "").replace(".", "").replace(" ", "").replace(",", ".")
    s = re.sub(r"[^0-9.\-]", "", s)
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _pct(v: Any) -> float | None:
    if v in (None, ""):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    m = re.search(r"(\d+(?:[.,]\d+)?)", _txt(v).replace("%", "").replace(" ", ""))
    return float(m.group(1).replace(",", ".")) if m else None


def _date(v: Any) -> str | None:
    if v in (None, ""):
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    for fmt in ("%d.%m.%y", "%d.%m.%Y", "%d/%m/%y", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(_txt(v), fmt).date().isoformat()
        except ValueError:
            pass
    return None


def _liq(v: Any) -> int | None:
    if v in (None, ""):
        return None
    if isinstance(v, (int, float)):
        return int(v)
    m = re.search(r"d\+?(\d+)", _txt(v).lower().replace(" ", ""))
    if m:
        return int(m.group(1))
    m = re.search(r"(\d+)", _txt(v))
    return int(m.group(1)) if m else None


def _institution(v: Any) -> str:
    s = _slug(v)
    for key, name in INSTITUTIONS.items():
        if key in s:
            return name
    return _txt(v)


def _indexer(*values: Any) -> str:
    s = _slug(" ".join(_txt(v) for v in values if _txt(v)))
    for key, name in INDEXER_MAP.items():
        if key in s:
            return name
    return ""


def _asset_class(value: Any, ticker: str = "", name: str = "") -> str:
    raw = _txt(value).upper()
    if raw in KNOWN_CLASSES:
        return raw
    s = _slug(f"{value} {ticker} {name}")
    if ticker:
        return "RV"
    if any(t in s for t in FUND_HINTS):
        return "FUNDOS"
    if any(t in s for t in RF_TOKENS):
        return "RF"
    return "OUTROS"


def _section_hint(sheet_name: str, row_values: list[Any], start_col: int) -> str:
    text = _slug(" ".join(_txt(v) for v in row_values[max(0, start_col - 3) : start_col + 1]))
    sheet = _slug(sheet_name)
    if "renda variavel" in text or "renda variavel" in sheet:
        return "RV"
    if "fund" in text or "fund" in sheet:
        return "FUNDOS"
    if "outro" in text or "outro" in sheet:
        return "OUTROS"
    if "renda fixa" in text or "renda fixa" in sheet:
        return "RF"
    return ""


def _header(v: Any) -> bool:
    return _slug(v) in {
        "tipo",
        "classe",
        "emissor",
        "instituicao",
        "taxa",
        "indexador",
        "vencimento",
        "valor",
        "liquidez",
        "ticker",
        "nome",
        "data de aquisicao",
        "data aquisicao",
    }


def _detect_sections(ws) -> list[dict[str, Any]]:
    sections: list[dict[str, Any]] = []
    for r in range(1, ws.max_row + 1):
        values = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        headers = [(c, _txt(values[c - 1])) for c in range(1, ws.max_column + 1) if _header(values[c - 1])]
        if len(headers) < 2:
            continue
        block = [headers[0]]
        for item in headers[1:]:
            if item[0] == block[-1][0] + 1:
                block.append(item)
                continue
            if len(block) >= 2:
                sections.append(
                    {
                        "row": r,
                        "start": block[0][0],
                        "end": block[-1][0],
                        "headers": [h for _, h in block],
                        "hint": _section_hint(ws.title, values, block[0][0] - 1),
                    }
                )
            block = [item]
        if len(block) >= 2:
            sections.append(
                {
                    "row": r,
                    "start": block[0][0],
                    "end": block[-1][0],
                    "headers": [h for _, h in block],
                    "hint": _section_hint(ws.title, values, block[0][0] - 1),
                }
            )
    return sections


def _pick(headers: list[str], *candidates: str) -> int | None:
    low = [_slug(h) for h in headers]
    for cand in candidates:
        cand = _slug(cand)
        for i, h in enumerate(low):
            if h == cand or cand in h:
                return i
    return None


def _parse_section(section: dict[str, Any], row_values: list[Any]) -> dict[str, Any] | None:
    headers = section["headers"]
    vals = row_values[section["start"] - 1 : section["start"] - 1 + len(headers)]
    if not any(v not in (None, "") for v in vals):
        return None

    slugged = [_slug(h) for h in headers]
    if "ticker" in slugged:
        i_ticker = _pick(headers, "ticker")
        i_val = _pick(headers, "valor")
        i_name = _pick(headers, "nome")
        if i_ticker is None or i_val is None:
            return None
        ticker = _txt(vals[i_ticker])
        if not ticker:
            return None
        issuer = _txt(vals[i_name]) if i_name is not None else ""
        return {
            "asset_class": "RV",
            "institution": issuer,
            "issuer": issuer or ticker,
            "ticker": ticker,
            "name": ticker,
            "indexer": "",
            "rate": None,
            "maturity": None,
            "current_value": _money(vals[i_val]),
            "liquidity_days": None,
            "acquisition_date": None,
            "cost_value": None,
            "source": "upload",
            "section": section["hint"] or "RV",
        }

    if any(h in {"tipo", "classe", "emissor", "taxa", "vencimento"} for h in slugged):
        i_tipo = _pick(headers, "tipo", "classe")
        i_emissor = _pick(headers, "emissor", "instituicao", "instituição")
        i_taxa = _pick(headers, "taxa", "indexador")
        i_venc = _pick(headers, "vencimento")
        i_val = _pick(headers, "valor")
        i_liq = _pick(headers, "liquidez")
        i_nome = _pick(headers, "nome")
        i_acq = _pick(headers, "data de aquisição", "data aquisicao", "aquisição")
        name = _txt(vals[i_nome]) if i_nome is not None else ""
        asset_class = _asset_class(vals[i_tipo] if i_tipo is not None else "", name=name)
        rate_text = _txt(vals[i_taxa]) if i_taxa is not None else ""
        return {
            "asset_class": asset_class,
            "institution": _txt(vals[i_emissor]) if i_emissor is not None else "",
            "issuer": _txt(vals[i_emissor]) if i_emissor is not None else "",
            "ticker": "",
            "name": name or _txt(vals[i_emissor]),
            "indexer": _indexer(rate_text, name),
            "rate": _pct(rate_text),
            "maturity": _date(vals[i_venc]) if i_venc is not None else None,
            "current_value": _money(vals[i_val]) if i_val is not None else None,
            "liquidity_days": _liq(vals[i_liq]) if i_liq is not None else None,
            "acquisition_date": _date(vals[i_acq]) if i_acq is not None else None,
            "cost_value": None,
            "source": "upload",
            "section": section["hint"] or asset_class,
        }

    if "nome" in slugged:
        i_nome = _pick(headers, "nome")
        i_val = _pick(headers, "valor")
        i_liq = _pick(headers, "liquidez")
        if i_nome is None or i_val is None:
            return None
        name = _txt(vals[i_nome])
        value = _money(vals[i_val])
        if not name or value is None:
            return None
        return {
            "asset_class": "FUNDOS" if any(t in _slug(name) for t in FUND_HINTS) else "OUTROS",
            "institution": "",
            "issuer": name,
            "ticker": "",
            "name": name,
            "indexer": "",
            "rate": None,
            "maturity": None,
            "current_value": value,
            "liquidity_days": _liq(vals[i_liq]) if i_liq is not None else None,
            "acquisition_date": None,
            "cost_value": None,
            "source": "upload",
            "section": section["hint"] or "OUTROS",
        }
    return None


def parse_xlsx(path: Path) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    rows: list[dict[str, Any]] = []
    for ws in wb.worksheets:
        for sec in _detect_sections(ws):
            for r in range(sec["row"] + 1, ws.max_row + 1):
                values = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
                parsed = _parse_section(sec, values)
                if parsed:
                    rows.append(parsed)
    return rows


def parse_csv(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        sample = fh.read(4096)
        fh.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample)
        except csv.Error:
            dialect = csv.excel
        return [normalize_row(row) for row in csv.DictReader(fh, dialect=dialect)]


def normalize_row(row: dict[str, Any]) -> dict[str, Any]:
    lower = {_slug(k): v for k, v in row.items() if k is not None}
    ticker = _txt(lower.get("ticker"))
    name = _txt(lower.get("nome") or lower.get("name") or ticker)
    asset_class = _asset_class(lower.get("asset class") or lower.get("classe") or lower.get("tipo") or lower.get("asset_class") or "", ticker=ticker, name=name)
    issuer = _institution(lower.get("emissor") or lower.get("issuer") or lower.get("instituicao") or lower.get("instituição") or "")
    return {
        "asset_class": asset_class,
        "institution": _institution(lower.get("institution") or lower.get("instituicao") or issuer),
        "issuer": _txt(lower.get("emissor") or lower.get("issuer") or issuer),
        "ticker": ticker,
        "name": name,
        "indexer": _txt(lower.get("indexer") or lower.get("indexador") or _indexer(lower.get("taxa"), lower.get("rate"))),
        "rate": _pct(lower.get("taxa") or lower.get("rate")),
        "maturity": _date(lower.get("vencimento") or lower.get("maturity")),
        "current_value": _money(lower.get("valor") or lower.get("current value") or lower.get("current_value") or lower.get("valor atual")),
        "liquidity_days": _liq(lower.get("liquidez") or lower.get("liquidity") or lower.get("liquidity_days")),
        "acquisition_date": _date(lower.get("data de aquisicao") or lower.get("data aquisicao") or lower.get("acquisition date") or lower.get("acquisition_date")),
        "cost_value": _money(lower.get("custo") or lower.get("valor aquisicao") or lower.get("cost value") or lower.get("cost_value")),
        "source": "upload",
        "section": asset_class,
    }


def normalize_holdings(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    holdings: list[dict[str, Any]] = []
    for idx, row in enumerate(rows, start=1):
        n = dict(row)
        n["id"] = f"h{idx}"
        n["asset_class"] = n.get("asset_class") or _asset_class("", ticker=_txt(n.get("ticker")), name=_txt(n.get("name")))
        n["institution"] = _institution(n.get("institution") or n.get("issuer") or "")
        n["issuer"] = _txt(n.get("issuer") or n.get("institution"))
        n["ticker"] = _txt(n.get("ticker"))
        n["name"] = _txt(n.get("name") or n.get("ticker") or n.get("issuer"))
        n["indexer"] = _txt(n.get("indexer") or "")
        n["rate"] = n.get("rate") if n.get("rate") is not None else _pct(n.get("taxa"))
        n["maturity"] = n.get("maturity") or _date(n.get("vencimento"))
        n["current_value"] = float(n.get("current_value") or 0.0)
        n["liquidity_days"] = n.get("liquidity_days")
        n["acquisition_date"] = n.get("acquisition_date")
        n["cost_value"] = n.get("cost_value")
        n["asset_class"] = _asset_class(n["asset_class"], ticker=n["ticker"], name=n["name"])
        holdings.append(n)
    return holdings


def series(holdings: list[dict[str, Any]], key: str, only: str | None = None) -> list[dict[str, Any]]:
    buckets: dict[str, float] = {}
    for row in holdings:
        if only and row["asset_class"] != only:
            continue
        label = row.get(key) or "Nao informado"
        buckets[label] = buckets.get(label, 0.0) + float(row.get("current_value") or 0.0)
    total = sum(buckets.values()) or 1.0
    return [{"label": k, "value": v, "share": v / total} for k, v in sorted(buckets.items(), key=lambda x: x[1], reverse=True)]


def maturity_series(holdings: list[dict[str, Any]]) -> list[dict[str, Any]]:
    order = ["30 dias", "6 meses", "1 ano", "3 anos", "3+ anos", "Sem vencimento"]
    buckets = {k: 0.0 for k in order}
    today = date.today()
    for row in holdings:
        if row["asset_class"] != "RF":
            continue
        if not row.get("maturity"):
            buckets["Sem vencimento"] += float(row.get("current_value") or 0.0)
            continue
        try:
            diff = (date.fromisoformat(row["maturity"]) - today).days
        except ValueError:
            buckets["Sem vencimento"] += float(row.get("current_value") or 0.0)
            continue
        label = "30 dias" if diff <= 30 else "6 meses" if diff <= 180 else "1 ano" if diff <= 365 else "3 anos" if diff <= 1095 else "3+ anos"
        buckets[label] += float(row.get("current_value") or 0.0)
    return [{"label": k, "value": buckets[k]} for k in order]


def indexer_series(holdings: list[dict[str, Any]]) -> list[dict[str, Any]]:
    buckets = {"CDI": 0.0, "IPCA": 0.0, "Pré": 0.0, "Outros": 0.0}
    for row in holdings:
        if row["asset_class"] != "RF":
            continue
        value = float(row.get("current_value") or 0.0)
        idx = (row.get("indexer") or "").upper()
        if "CDI" in idx:
            buckets["CDI"] += value
        elif "IPCA" in idx:
            buckets["IPCA"] += value
        elif "PR" in idx:
            buckets["Pré"] += value
        else:
            buckets["Outros"] += value
    total = sum(buckets.values()) or 1.0
    return [{"label": k, "value": v, "share": v / total} for k, v in buckets.items()]


def benchmark_demo(holdings: list[dict[str, Any]]) -> dict[str, Any]:
    total = sum(float(r.get("current_value") or 0.0) for r in holdings) or 1.0
    rf = sum(float(r.get("current_value") or 0.0) for r in holdings if r["asset_class"] == "RF")
    rv = sum(float(r.get("current_value") or 0.0) for r in holdings if r["asset_class"] == "RV")
    fund = sum(float(r.get("current_value") or 0.0) for r in holdings if r["asset_class"] == "FUNDOS")
    comp = [
        {"label": "CDI", "value": rf * 0.68, "share": (rf * 0.68) / total},
        {"label": "IPCA", "value": rf * 0.2, "share": (rf * 0.2) / total},
        {"label": "Pré", "value": rf * 0.12, "share": (rf * 0.12) / total},
        {"label": "IBOV", "value": rv + fund * 0.35, "share": (rv + fund * 0.35) / total},
    ]
    data = []
    cdi = ibov = ipca = port = 100.0
    for label, step in zip(
        ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"],
        [
            (0.0075, 0.012, 0.0042, 0.0091),
            (0.0074, 0.010, 0.0044, 0.0088),
            (0.0076, 0.011, 0.0041, 0.0089),
            (0.0073, 0.009, 0.0043, 0.0085),
            (0.0077, 0.013, 0.0045, 0.0094),
            (0.0075, 0.008, 0.0040, 0.0087),
            (0.0076, 0.011, 0.0042, 0.0090),
            (0.0075, 0.012, 0.0041, 0.0092),
            (0.0074, 0.010, 0.0043, 0.0088),
            (0.0076, 0.009, 0.0042, 0.0089),
            (0.0075, 0.011, 0.0044, 0.0091),
            (0.0077, 0.013, 0.0045, 0.0095),
        ],
    ):
        cdi *= 1 + step[0]
        ibov *= 1 + step[1]
        ipca *= 1 + step[2]
        port *= 1 + step[3]
        data.append({"label": label, "CDI": round(cdi, 2), "IBOV": round(ibov, 2), "IPCA": round(ipca, 2), "Carteira": round(port, 2)})
    return {"composition": comp, "series": data}


def correlation_demo(holdings: list[dict[str, Any]]) -> dict[str, Any]:
    sample = holdings[:8]
    labels = [(_txt(r.get("ticker")) or _txt(r.get("issuer")) or _txt(r.get("name")))[:18] for r in sample]
    matrix: list[list[float]] = []
    for a in sample:
        line: list[float] = []
        for b in sample:
            score = 0.08
            if a["asset_class"] == b["asset_class"]:
                score += 0.35
            if (a.get("indexer") or "") == (b.get("indexer") or "") and a.get("indexer"):
                score += 0.25
            if _slug(a.get("issuer")) == _slug(b.get("issuer")) and a.get("issuer"):
                score += 0.22
            if _slug(a.get("institution")) == _slug(b.get("institution")) and a.get("institution"):
                score += 0.12
            if a.get("liquidity_days") and b.get("liquidity_days") and abs(int(a["liquidity_days"]) - int(b["liquidity_days"])) <= 5:
                score += 0.05
            line.append(round(min(score, 0.98), 2))
        matrix.append(line)
    return {"labels": labels, "matrix": matrix}


def insights(holdings: list[dict[str, Any]], by_issuer: list[dict[str, Any]], by_indexer: list[dict[str, Any]]) -> list[str]:
    total = sum(float(r.get("current_value") or 0.0) for r in holdings) or 1.0
    out: list[str] = []
    if by_issuer:
        top = by_issuer[0]
        if top["share"] >= 0.2:
            out.append(f"Alta exposicao ao emissor {top['label']} ({top['share']*100:.1f}% da carteira).")
    if by_indexer:
        top = by_indexer[0]
        if top["label"] == "CDI" and top["share"] >= 0.7:
            out.append("Excesso em CDI pode reduzir ganho real em cenarios de inflacao persistente.")
        elif top["label"] == "Pré" and top["share"] >= 0.7:
            out.append("Carteira muito sensivel a marcação a mercado por excesso de prefixados.")
    illiquid = sum(float(r.get("current_value") or 0.0) for r in holdings if (r.get("liquidity_days") or 0) >= 30)
    if illiquid / total >= 0.15:
        out.append("Liquidez concentrada em ativos com D+30 ou superior.")
    if len([r for r in by_issuer if r["value"] > 0]) <= 8:
        out.append("Base de emissores enxuta; vale observar risco de falsa diversificacao.")
    if len({r["asset_class"] for r in holdings}) <= 2:
        out.append("Carteira conservadora com baixa diversificacao entre classes.")
    return out[:6] or ["Carteira com distribuicao equilibrada dentro da amostra atual."]


def score_portfolio(holdings: list[dict[str, Any]], by_issuer: list[dict[str, Any]], by_indexer: list[dict[str, Any]]) -> int:
    total = sum(float(r.get("current_value") or 0.0) for r in holdings) or 1.0
    score = 100
    if by_issuer:
        score -= int(min(25, by_issuer[0]["share"] * 60))
    if by_indexer:
        score -= int(min(15, by_indexer[0]["share"] * 35))
    illiquid = sum(float(r.get("current_value") or 0.0) for r in holdings if (r.get("liquidity_days") or 0) >= 30)
    score -= int(min(10, (illiquid / total) * 25))
    if len({r["asset_class"] for r in holdings}) <= 2:
        score -= 8
    if len({r.get("issuer") for r in holdings if r.get("issuer")}) <= 5:
        score -= 8
    return max(0, min(100, score))


def compute_portfolio(rows: list[dict[str, Any]]) -> dict[str, Any]:
    holdings = normalize_holdings(rows)
    total = sum(float(r.get("current_value") or 0.0) for r in holdings)
    by_class = series(holdings, "asset_class")
    by_inst = series(holdings, "institution")
    by_issuer = series(holdings, "issuer")
    by_ticker = series(holdings, "ticker", only="RV")
    by_fund = series(holdings, "name", only="FUNDOS")
    by_indexer = indexer_series(holdings)
    maturity = maturity_series(holdings)
    bench = benchmark_demo(holdings)
    corr = correlation_demo(holdings)
    score = score_portfolio(holdings, by_issuer, by_indexer)
    alerts = []
    for row in by_issuer[:10]:
        if row["value"] > 250_000:
            alerts.append({"severity": "danger", "title": "FGC superado", "detail": f"Exposicao em {row['label']} acima de R$ 250 mil."})
        elif row["value"] > 200_000:
            alerts.append({"severity": "warning", "title": "Concentracao relevante", "detail": f"Exposicao em {row['label']} acima de R$ 200 mil."})
    if by_indexer and by_indexer[0]["share"] >= 0.7:
        alerts.append({"severity": "warning", "title": "Concentracao em indexador", "detail": f"{by_indexer[0]['label']} representa {by_indexer[0]['share']*100:.1f}% da carteira."})
    illiquid = sum(float(r.get("current_value") or 0.0) for r in holdings if (r.get("liquidity_days") or 0) >= 30) / (total or 1.0)
    if illiquid >= 0.15:
        alerts.append({"severity": "danger" if illiquid >= 0.3 else "warning", "title": "Liquidez ruim", "detail": f"{illiquid*100:.1f}% da carteira está em ativos com D+30 ou superior."})
    if by_issuer and by_issuer[0]["share"] >= 0.25:
        alerts.append({"severity": "warning", "title": "Diversificacao fraca", "detail": f"{by_issuer[0]['label']} ocupa {by_issuer[0]['share']*100:.1f}% do patrimonio."})
    if any(r["value"] > 0 and r["label"] == "30 dias" for r in maturity):
        alerts.append({"severity": "info", "title": "Vencimentos proximos", "detail": "Ha fluxo relevante de vencimentos nos proximos 30 dias."})
    return {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "summary": {
            "total": total,
            "holdings_count": len(holdings),
            "class_count": len({r["asset_class"] for r in holdings}),
            "issuer_count": len({r["issuer"] for r in holdings if r["issuer"]}),
            "score": score,
        },
        "holdings": holdings,
        "by_class": by_class,
        "by_institution": by_inst,
        "by_issuer": by_issuer,
        "by_indexer": by_indexer,
        "by_ticker": by_ticker,
        "by_fund": by_fund,
        "maturity": maturity,
        "benchmark": bench,
        "correlation": corr,
        "alerts": alerts,
        "insights": insights(holdings, by_issuer, by_indexer),
    }


DEMO = [
    {"asset_class": "RF", "institution": "BTG", "issuer": "BTG", "name": "BTG CDB", "indexer": "CDI", "rate": 100, "maturity": "2027-11-29", "current_value": 27179.80, "source": "demo"},
    {"asset_class": "RF", "institution": "BTG", "issuer": "BTG", "name": "BTG CDB", "indexer": "CDI", "rate": 100, "maturity": "2027-12-20", "current_value": 83126.03, "source": "demo"},
    {"asset_class": "RF", "institution": "Digimais", "issuer": "Digimais", "name": "Digimais CDB", "indexer": "CDI", "rate": 120, "maturity": "2028-02-18", "current_value": 26060.09, "source": "demo"},
    {"asset_class": "RF", "institution": "Digimais", "issuer": "Digimais", "name": "Digimais CDB", "indexer": "CDI", "rate": 121, "maturity": "2028-05-25", "current_value": 72465.38, "source": "demo"},
    {"asset_class": "RF", "institution": "Banco Arbi", "issuer": "Banco Arbi", "name": "Banco Arbi CDB", "indexer": "CDI", "rate": 114, "maturity": "2031-01-20", "current_value": 237121.40, "source": "demo"},
    {"asset_class": "RF", "institution": "BTG", "issuer": "BTG", "name": "BTG CRA", "indexer": "CDI", "rate": 1, "maturity": "2033-09-15", "current_value": 149751.27, "source": "demo"},
    {"asset_class": "RF", "institution": "Iguatemi", "issuer": "Iguatemi", "name": "Iguatemi CRI", "indexer": "CDI", "rate": 0.4, "maturity": "2034-09-21", "current_value": 92161.74, "source": "demo"},
    {"asset_class": "RF", "institution": "Banco do Nordeste", "issuer": "Banco do Nordeste", "name": "Banco do Nordeste LF", "indexer": "CDI", "rate": 165, "maturity": "2050-08-04", "current_value": 313232.84, "source": "demo"},
    {"asset_class": "RF", "institution": "BRB", "issuer": "BRB", "name": "BRB LF", "indexer": "CDI", "rate": 135, "maturity": "2028-10-19", "current_value": 519326.38, "source": "demo"},
    {"asset_class": "RF", "institution": "BMG", "issuer": "BMG", "name": "BMG CDB", "indexer": "CDI", "rate": 15.3, "maturity": "2026-06-08", "current_value": 142037.83, "liquidity_days": 30, "source": "demo"},
    {"asset_class": "RF", "institution": "PicPay", "issuer": "PicPay", "name": "PicPay CDB", "indexer": "CDI", "rate": 15.7, "maturity": "2028-10-30", "current_value": 19965.56, "liquidity_days": 15, "source": "demo"},
    {"asset_class": "RF", "institution": "Original", "issuer": "Original", "name": "Original CDB", "indexer": "CDI", "rate": 13.52, "maturity": "2027-06-21", "current_value": 210537.24, "liquidity_days": 15, "source": "demo"},
    {"asset_class": "RF", "institution": "Facta Financeira", "issuer": "Facta Financeira", "name": "Facta Financeira CDB", "indexer": "CDI", "rate": 16.8, "maturity": "2026-11-30", "current_value": 44101.61, "liquidity_days": 15, "source": "demo"},
    {"asset_class": "RF", "institution": "Semear", "issuer": "Semear", "name": "Semear CDB", "indexer": "CDI", "rate": 16.7, "maturity": "2027-04-15", "current_value": 65020.18, "liquidity_days": 15, "source": "demo"},
    {"asset_class": "RF", "institution": "Equatorial Goiás", "issuer": "Equatorial Goiás", "name": "Equatorial Goiás DEB", "indexer": "IPCA", "rate": 7.1, "maturity": "2031-04-15", "current_value": 212506.29, "liquidity_days": 30, "source": "demo"},
    {"asset_class": "RF", "institution": "Igua", "issuer": "Iguá", "name": "Igua DEB", "indexer": "IPCA", "rate": 7.4, "maturity": "2043-05-15", "current_value": 555597.94, "liquidity_days": 30, "source": "demo"},
    {"asset_class": "RV", "institution": "", "issuer": "BODB11", "ticker": "BODB11", "name": "BODB11", "current_value": 18879.38, "source": "demo"},
    {"asset_class": "RV", "institution": "", "issuer": "GZIT11", "ticker": "GZIT11", "name": "GZIT11", "current_value": 76248.00, "source": "demo"},
    {"asset_class": "RV", "institution": "", "issuer": "BDIF11", "ticker": "BDIF11", "name": "BDIF11", "current_value": 136952.66, "source": "demo"},
    {"asset_class": "RV", "institution": "", "issuer": "PLPL3", "ticker": "PLPL3", "name": "PLPL3", "current_value": 146948.86, "source": "demo"},
    {"asset_class": "FUNDOS", "institution": "BTG", "issuer": "BTG Pactual Infra", "name": "BTG Pactual Infra", "current_value": 101634.90, "liquidity_days": 10, "source": "demo"},
    {"asset_class": "FUNDOS", "institution": "Safra", "issuer": "Safra Infra conceito", "name": "Safra Infra conceito", "current_value": 918042.79, "liquidity_days": 30, "source": "demo"},
    {"asset_class": "FUNDOS", "institution": "Safra", "issuer": "Safra CAPMKT Infra", "name": "Safra CAPMKT Infra", "current_value": 501192.98, "liquidity_days": 30, "source": "demo"},
]


def _load_state() -> list[dict[str, Any]]:
    _ensure_storage()
    if STATE_FILE.exists():
        try:
            payload = json.loads(STATE_FILE.read_text(encoding="utf-8"))
            if isinstance(payload, list):
                return payload
        except Exception:
            pass
    return DEMO


def _save_state(rows: list[dict[str, Any]]) -> None:
    _ensure_storage()
    STATE_FILE.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")


app.state.rows = _load_state()


@app.get("/", response_class=HTMLResponse)
def home() -> HTMLResponse:
    return HTMLResponse((STATIC_DIR / "index.html").read_text(encoding="utf-8"))


@app.get("/api/state")
def state() -> JSONResponse:
    return JSONResponse(compute_portfolio(app.state.rows))


@app.post("/api/reset-demo")
def reset_demo() -> JSONResponse:
    app.state.rows = DEMO
    _save_state(app.state.rows)
    return JSONResponse({"ok": True, "state": compute_portfolio(app.state.rows)})


@app.post("/api/manual")
async def manual(row: dict[str, Any]) -> JSONResponse:
    if not isinstance(row, dict):
        raise HTTPException(status_code=400, detail="Corpo invalido")
    rows = list(app.state.rows)
    row_id = _txt(row.get("id"))
    normalized = normalize_row(row)
    normalized["source"] = "manual"
    if row_id and row_id.startswith("h") and row_id[1:].isdigit():
        idx = int(row_id[1:]) - 1
        if 0 <= idx < len(rows):
            rows[idx] = normalized
        else:
            rows.append(normalized)
    else:
        rows.append(normalized)
    app.state.rows = rows
    _save_state(app.state.rows)
    return JSONResponse({"ok": True, "state": compute_portfolio(app.state.rows)})


@app.post("/api/delete")
async def delete(row: dict[str, Any]) -> JSONResponse:
    if not isinstance(row, dict):
        raise HTTPException(status_code=400, detail="Corpo invalido")
    row_id = _txt(row.get("id"))
    if not row_id.startswith("h") or not row_id[1:].isdigit():
        raise HTTPException(status_code=400, detail="ID invalido")
    idx = int(row_id[1:]) - 1
    rows = list(app.state.rows)
    if idx < 0 or idx >= len(rows):
        raise HTTPException(status_code=404, detail="Registro nao encontrado")
    rows.pop(idx)
    app.state.rows = rows
    _save_state(app.state.rows)
    return JSONResponse({"ok": True, "state": compute_portfolio(app.state.rows)})


@app.post("/api/upload")
async def upload(file: UploadFile = File(...)) -> JSONResponse:
    name = (file.filename or "").lower()
    suffix = Path(name).suffix
    temp_dir = STORAGE_DIR / "uploads"
    temp_dir.mkdir(exist_ok=True)
    temp = temp_dir / f"upload_{int(datetime.now().timestamp())}{suffix}"
    temp.write_bytes(await file.read())
    try:
        if suffix in {".xlsx", ".xlsm", ".xls"}:
            rows = parse_xlsx(temp)
        elif suffix in {".csv", ".txt"}:
            rows = parse_csv(temp)
        else:
            raise HTTPException(status_code=400, detail="Formato nao suportado. Use XLSX ou CSV.")
        if not rows:
            raise HTTPException(status_code=400, detail="Nenhum dado reconhecido na planilha.")
        app.state.rows = rows
        _save_state(app.state.rows)
        return JSONResponse({"ok": True, "state": compute_portfolio(app.state.rows)})
    finally:
        try:
            temp.unlink(missing_ok=True)
        except Exception:
            pass


if __name__ == "__main__":
    import uvicorn

    uvicorn.run("main:app", host="127.0.0.1", port=8000, reload=True)
